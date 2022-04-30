from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook,load_workbook
from matplotlib import pyplot as plt
import networkx as nx
import numpy as np

from kivy.config import Config
Config.set('input', 'mouse', 'mouse,disable_multitouch')
Config.set('kivy', 'exit_on_escape', '0')
Config.set('graphics','resizable', '0')

from kivy.lang import Builder
Builder.load_file('Components/GraphWriting.kv')
Builder.load_file('Components/PreSteps.kv')
Builder.load_file('Components/MainMenu.kv')
Builder.load_file('Components/Results.kv')
Builder.load_file('Components/Widgets.kv')

from kivy.uix.screenmanager import Screen,ScreenManager
from kivy.garden.matplotlib import FigureCanvasKivyAgg
from kivy.uix.screenmanager import NoTransition
from kivymd.uix.gridlayout import MDGridLayout
from kivy.uix.floatlayout import FloatLayout
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.list import OneLineListItem
from tkinter import filedialog,Tk,StringVar
from kivy.core.window import Window
from kivymd.uix.card import MDCard
from kivymd.app import MDApp

class WindowManager(ScreenManager): pass
class GraphCreationScreen(Screen): pass
class MainMenuScreen(Screen): pass
class SolvedScreen(Screen): pass
class PreUpScreen(Screen): pass
class DoneScreen(Screen): pass
class LowerContent(MDGridLayout): pass
class VistaScroll(FloatLayout): pass
class Content(MDBoxLayout): pass
class NombreNodoSource(MDCard): pass
class AddButton(MDCard): pass

def write_excel(nombre,cant_nodos,pares):
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    if '.xlsx' not in nombre:
        nombre += '.xlsx'
    book = Workbook()
    sheet = book.active
    for fila in range(1,cant_nodos+1):
        sheet.cell(row=fila+1, column=1).value = fila
        sheet.cell(row=fila+1, column=1).fill = PatternFill("solid", start_color="5cb800")
        sheet.cell(row=fila+1, column=1).alignment = Alignment(horizontal='center')
        sheet.cell(row=fila+1, column=1).border = thin_border
    for columna in range(1,cant_nodos+1):
        sheet.cell(row=1, column=columna+1).value = columna
        sheet.cell(row=1, column=columna+1).fill = PatternFill("solid", start_color="5cb800")
        sheet.cell(row=1, column=columna+1).alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=columna+1).border = thin_border   
    for posicion in pares:
        sheet.cell(row=int(posicion[0])+1, column=int(posicion[1])+1).value = int(pares[posicion])
        sheet.cell(row=int(posicion[0])+1, column=int(posicion[1])+1).alignment = Alignment(horizontal='center')
        sheet.cell(row=int(posicion[0])+1, column=int(posicion[1])+1).border = thin_border
        sheet.cell(row=int(posicion[1])+1, column=int(posicion[0])+1).value = int(pares[posicion])
        sheet.cell(row=int(posicion[1])+1, column=int(posicion[0])+1).alignment = Alignment(horizontal='center')
        sheet.cell(row=int(posicion[1])+1, column=int(posicion[0])+1).border = thin_border
    book.save(nombre)
    print('Matriz creada exitosamente\n')

def read_excel(nombre_excel):
    try:
        book = load_workbook(filename= nombre_excel)
        sheet = book.active
        A = np.array([[i.value if i.value is not None else 0 for i in j[1::]] for j in sheet.rows])
        A = np.delete(A,0,0)
        if  np.array_equal(np.transpose(A),A):
            print('Grafo no dirigido encontrado...')
            Grafo = nx.from_numpy_matrix(A)
        else:
            print('Grafo dirigido encontrado...')
            Grafo = nx.from_numpy_matrix(A, create_using=nx.DiGraph)
        print('Matriz importada exitosamente\n')
        return Grafo
    except Exception as e:
        print(e)
        return None

class Network_Manager(MDApp):
    def __init__(self,**kwargs):
        super().__init__(**kwargs)
        self.sm = WindowManager(transition=NoTransition())
        self.screenShelve={}
        self.screenShelve['MainMenu']=MainMenuScreen(name='Main')
        self.screenShelve['Pre']=PreUpScreen(name='Presteps')
        self.screenShelve['Solution']=SolvedScreen(name='Solve')
        self.screenShelve['Create']=GraphCreationScreen(name='Creation')
        self.screenShelve['Done']=DoneScreen(name='Done')  
        for i in self.screenShelve: self.sm.add_widget(self.screenShelve[i])
    def return_home(self):
        try:
            App.screenShelve['Pre'].ids.extaButton.children[0].ids.sourceName.text = ''
        except:
            pass
        App.screenShelve['Pre'].ids.pathExcel.text = ''
        self.last_seleccion = ''
        self.sm.current = 'Main'
        self.elementos = {}
        self.nodos=set()
        plt.clf()
    def on_start(self, **kwargs):
        self.return_home()
        
    def confirmacionAlgoritmo(self):
        nombre = App.screenShelve['Pre'].ids.pathExcel.text
        Grafo = read_excel(nombre)
        if Grafo is not None:
            if self.last_seleccion == '2':
                self.dijkstra(Grafo)
                self.sm.current = 'Solve'
            elif self.last_seleccion == '3':
                self.kruskal(Grafo)
                self.sm.current = 'Solve'
        else:
            App.screenShelve['Pre'].ids.pathExcel.text = 'ERROR'
    def confirmacionCreacion(self):
        excel_file_name = App.screenShelve['Create'].ids.nombreExcel.text
        if len(self.nodos)>= 6:
            if len(self.elementos)>=10:
                write_excel(excel_file_name,len(self.nodos),self.elementos)
                self.sm.current = 'Done'
    def confirmName(self):
            new_name = App.screenShelve['Create'].ids.nombreSubida.text
            data = new_name.split(',')

            if new_name != '' and (data[0],data[1]) not in self.elementos and data[2].isnumeric():
                App.screenShelve['Create'].ids.FoundList.add_widget(OneLineListItem(text=f'({new_name})'))
                self.elementos[(data[0],data[1])] = int(data[2])
                self.nodos.add(data[0])
                self.nodos.add(data[1])



    def main(self,seleccion):
        #print('\n\nSeleccione una opción:\n[1] Crear parametros\n[2] Aplicar Dijkstra\n[3] Aplicar Kruskal\n[4] Salir\n')
        self.last_seleccion = seleccion
        if seleccion == '1':
            #nodos,pares = self.parametros()
            #write_excel(nodos,pares) 
            self.sm.current = 'Creation'
        elif seleccion == '2':
            App.screenShelve['Pre'].ids.extaButton.add_widget(NombreNodoSource())
            self.sm.current = 'Presteps'        
        elif seleccion == '3':
            widgets,nombres= App.screenShelve['Pre'].ids.extaButton.children, []
            for i in widgets: nombres.append(i)
            for i in nombres:  App.screenShelve['Pre'].ids.extaButton.remove_widget(i)
            self.sm.current = 'Presteps'
            #nombre = input('Ingrese nombre archivo excel: ')
        elif seleccion == '4':
            MDApp.get_running_app().stop()
            Window.close()
            print('Programa finalizado.')
        else:
            print('Opción ingresada no valida')

    def dijkstra(self,grafo):
        source_node  = int(App.screenShelve['Pre'].ids.extaButton.children[0].ids.sourceName.text)#int(input('Ingrese el nodo inicial: '))
        data,_ =nx.dijkstra_predecessor_and_distance(grafo,source_node)
        
        shortest_edges = [(node, data[node][0]) if data[node]!=[] else (None,None) for node in data]
        shortest_edges.remove((None,None))
       
        temp_labels = nx.get_edge_attributes(grafo,'weight')
        G = nx.Graph()
        for arista in shortest_edges:
            if arista in temp_labels:
                G.add_edge(arista[0], arista[1], weight=temp_labels[arista])
            if (arista[1],arista[0]) in temp_labels:
                G.add_edge(arista[1], arista[0], weight=temp_labels[(arista[1],arista[0])])
        pos=nx.spring_layout(G) # 
        nx.draw(G,pos,node_color='#73ABD0',node_size=450,with_labels=True)
        labels = nx.get_edge_attributes(G,'weight')
        nx.draw_networkx_edge_labels(G,pos,edge_labels=labels)
        App.screenShelve['Solution'].ids.GrafoImagen.add_widget(FigureCanvasKivyAgg(plt.gcf()))        

    def kruskal(self, grafo):
        G = nx.minimum_spanning_tree(grafo,algorithm='kruskal')
        pos=nx.spring_layout(G) # 
        nx.draw(G,pos,node_color='#73ABD0',node_size=450,with_labels=True)
        labels = nx.get_edge_attributes(G,'weight')
        nx.draw_networkx_edge_labels(G,pos,edge_labels=labels)
        App.screenShelve['Solution'].ids.GrafoImagen.add_widget(FigureCanvasKivyAgg(plt.gcf()))

    def _on_file_drop(self, window, file_path):
        App.screenShelve['Pre'].ids.pathExcel.text = file_path.decode()
        App.screenShelve['Pre'].ids.pathExcel.cursor = (0,0)

    def build(self):
        self.title = 'Network Assigment 1'
#        self.icon=f"{self.resources_path}/globe256.png"
        Window.bind(on_dropfile=self._on_file_drop)
        return self.sm
if __name__ == '__main__':
    App = Network_Manager()
    App.run()